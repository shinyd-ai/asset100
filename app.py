from flask import Flask, render_template, request, jsonify, Response, g
from flask_cors import CORS
from database import get_db as database_get_db, init_db
from datetime import datetime, date
import json, os, shutil, sqlite3, re, csv, io, requests
import yfinance as yf

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

app = Flask(__name__)
CORS(app)

def get_db():
    if 'db' not in g:
        g.db = database_get_db()
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db is not None:
        db.close()

# ── 페이지 라우터 ────────────────────────────────────────────
@app.route('/')
def dashboard():
    return render_template('dashboard.html')

@app.route('/income')
def income():
    return render_template('income.html')

@app.route('/budget')
def budget():
    return render_template('budget.html')

@app.route('/cards')
def cards():
    return render_template('cards.html')

@app.route('/investments')
def investments():
    return render_template('investments.html')

@app.route('/realestate')
def realestate():
    return render_template('realestate.html')

@app.route('/loans')
def loans():
    return render_template('loans.html')

@app.route('/pension')
def pension():
    return render_template('pension.html')

@app.route('/goals')
def goals():
    return render_template('goals.html')

@app.route('/monthly')
def monthly():
    return render_template('monthly.html')


# ── 공통 헬퍼 ────────────────────────────────────────────────
def rows_to_list(rows):
    return [dict(r) for r in rows]


# ── API: 수입 ────────────────────────────────────────────────
@app.route('/api/income', methods=['GET', 'POST'])
def api_income():
    db = get_db()
    if request.method == 'GET':
        year  = request.args.get('year')
        month = request.args.get('month')
        query = "SELECT * FROM income"
        params = []
        if year and month:
            query += " WHERE strftime('%Y', date) = ? AND strftime('%m', date) = ?"
            params = [year, month.zfill(2)]
        query += " ORDER BY date DESC"
        rows = db.execute(query, params).fetchall()
        return jsonify(rows_to_list(rows))

    data = request.json
    db.execute(
        "INSERT INTO income (date, category, name, memo, amount) VALUES (?,?,?,?,?)",
        (data['date'], data.get('category'), data.get('name'), data.get('memo'), data['amount'])
    )
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/income/<int:rid>', methods=['PUT', 'DELETE'])
def api_income_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        db.execute(
            "UPDATE income SET date=?, category=?, name=?, memo=?, amount=? WHERE id=?",
            (data.get('date'), data.get('category'), data.get('name'),
             data.get('memo'), data.get('amount', 0), rid)
        )
        db.commit()
        return jsonify({'ok': True})
    db.execute("DELETE FROM income WHERE id = ?", (rid,))
    db.commit()
    return jsonify({'ok': True})


# ── API: 가계부 ──────────────────────────────────────────────
def _sync_card_tx(db, budget_id, data):
    """budget 저장 시 card_tx 자동 동기화"""
    card_id = data.get('card_id') or None
    if card_id:
        existing = db.execute("SELECT id FROM card_tx WHERE budget_id = ?", (budget_id,)).fetchone()
        if existing:
            db.execute(
                "UPDATE card_tx SET card_id=?, date=?, name=?, category=?, amount=?, memo=? WHERE budget_id=?",
                (card_id, data.get('date'), data.get('name'), data.get('category'),
                 data.get('amount', 0), data.get('memo'), budget_id)
            )
        else:
            db.execute(
                "INSERT INTO card_tx (card_id, date, name, category, amount, installment, memo, budget_id) VALUES (?,?,?,?,?,1,?,?)",
                (card_id, data.get('date'), data.get('name'), data.get('category'),
                 data.get('amount', 0), data.get('memo'), budget_id)
            )
    else:
        db.execute("DELETE FROM card_tx WHERE budget_id = ?", (budget_id,))


@app.route('/api/budget', methods=['GET', 'POST'])
def api_budget():
    db = get_db()
    if request.method == 'GET':
        year  = request.args.get('year')
        month = request.args.get('month')
        query = """SELECT b.*, c.card_name
                   FROM budget b
                   LEFT JOIN card_info c ON b.card_id = c.id"""
        params = []
        if year and month:
            query += " WHERE strftime('%Y', b.date) = ? AND strftime('%m', b.date) = ?"
            params = [year, month.zfill(2)]
        query += " ORDER BY b.date DESC"
        rows = db.execute(query, params).fetchall()
        return jsonify(rows_to_list(rows))

    data = request.json
    card_id = data.get('card_id')
    if not card_id or str(card_id) == 'None' or str(card_id).strip() == '':
        card_id = None
        
    row = db.execute(
        "INSERT INTO budget (date, category, name, type, payment_method, amount, memo, card_id) VALUES (?,?,?,?,?,?,?,?) RETURNING id",
        (data['date'], data.get('category'), data.get('name'), data.get('type'),
         data.get('payment_method'), data['amount'], data.get('memo'), card_id)
    ).fetchone()
    budget_id = row['id']
    _sync_card_tx(db, budget_id, data)
    db.commit()
    return jsonify({'ok': True}), 201


@app.route('/api/budget/<int:rid>', methods=['PUT', 'DELETE'])
def api_budget_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        card_id = data.get('card_id')
        if not card_id or str(card_id) == 'None' or str(card_id).strip() == '':
            card_id = None
            
        db.execute(
            "UPDATE budget SET date=?, category=?, name=?, type=?, payment_method=?, amount=?, memo=?, card_id=? WHERE id=?",
            (data.get('date'), data.get('category'), data.get('name'), data.get('type'),
             data.get('payment_method'), data.get('amount', 0), data.get('memo'),
             card_id, rid)
        )
        _sync_card_tx(db, rid, data)
        db.commit()
        return jsonify({'ok': True})
    db.execute("DELETE FROM card_tx WHERE budget_id = ?", (rid,))
    db.execute("DELETE FROM budget WHERE id = ?", (rid,))
    db.commit()
    return jsonify({'ok': True})


# ── API: 카드 정보 ───────────────────────────────────────────
@app.route('/api/cards', methods=['GET', 'POST'])
def api_cards():
    db = get_db()
    if request.method == 'GET':
        rows = db.execute("SELECT * FROM card_info ORDER BY card_num").fetchall()
        return jsonify(rows_to_list(rows))

    data = request.json
    db.execute(
        "INSERT INTO card_info (card_num, card_name, limit_amount, payment_day, billing_day, benefit) VALUES (?,?,?,?,?,?)",
        (data['card_num'], data.get('card_name'), data.get('limit_amount', 0),
         data.get('payment_day'), data.get('billing_day'), data.get('benefit'))
    )
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/cards/<int:cid>', methods=['PUT', 'DELETE'])
def api_card_detail(cid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        db.execute(
            "UPDATE card_info SET card_num=?, card_name=?, limit_amount=?, payment_day=?, billing_day=?, benefit=? WHERE id=?",
            (data.get('card_num'), data.get('card_name'), data.get('limit_amount', 0),
             data.get('payment_day'), data.get('billing_day'), data.get('benefit'), cid)
        )
        db.commit()
        return jsonify({'ok': True})
    db.execute("UPDATE card_tx SET card_id = NULL WHERE card_id = ?", (cid,))
    db.execute("UPDATE budget SET card_id = NULL WHERE card_id = ?", (cid,))
    db.execute("DELETE FROM card_mappings WHERE card_id = ?", (cid,))
    db.execute("DELETE FROM card_info WHERE id = ?", (cid,))
    db.commit()
    return jsonify({'ok': True})


# ── API: 카드 거래내역 ───────────────────────────────────────
@app.route('/api/card-tx', methods=['GET', 'POST'])
def api_card_tx():
    db = get_db()
    if request.method == 'GET':
        card_id = request.args.get('card_id')
        year    = request.args.get('year')
        month   = request.args.get('month')
        query   = "SELECT t.*, c.card_name FROM card_tx t LEFT JOIN card_info c ON t.card_id = c.id"
        params  = []
        conds   = []
        if card_id:
            conds.append("t.card_id = ?"); params.append(card_id)
        if year and month:
            conds.append("strftime('%Y', t.date) = ?"); params.append(year)
            conds.append("strftime('%m', t.date) = ?"); params.append(month.zfill(2))
        if conds:
            query += " WHERE " + " AND ".join(conds)
        query += " ORDER BY t.date DESC"
        rows = db.execute(query, params).fetchall()
        total = sum(r['amount'] for r in rows)
        # 카테고리별 집계 (없는 건 → '-')
        cat_map = {}
        for r in rows:
            cat = r['category'] or '-'
            cat_map[cat] = cat_map.get(cat, 0) + r['amount']
        by_category = sorted(
            [{'category': c, 'total': t} for c, t in cat_map.items()],
            key=lambda x: x['total'], reverse=True
        )
        # 자금 그룹별 집계
        fund_group_names = {r['id']: r['name'] for r in db.execute("SELECT id, name FROM fund_groups").fetchall()}
        fund_map = {}
        for r in rows:
            gid = r['fund_group_id']
            name = fund_group_names.get(gid, '-') if gid else '-'
            fund_map[name] = fund_map.get(name, 0) + r['amount']
        by_fund_group = sorted(
            [{'name': n, 'total': t} for n, t in fund_map.items()],
            key=lambda x: x['total'], reverse=True
        )
        return jsonify({'rows': rows_to_list(rows), 'total': total,
                        'by_category': by_category, 'by_fund_group': by_fund_group})

    data = request.json
    db.execute(
        "INSERT INTO card_tx (card_id, date, name, category, amount, installment, memo) VALUES (?,?,?,?,?,?,?)",
        (data.get('card_id'), data['date'], data.get('name'), data.get('category'),
         data['amount'], data.get('installment', 1), data.get('memo'))
    )
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/card-tx/<int:rid>', methods=['PUT', 'DELETE'])
def api_card_tx_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        category = data.get('category') or ''
        locked = 1 if category else 0
        
        card_id = data.get('card_id')
        if not card_id or str(card_id) == 'None' or str(card_id).strip() == '':
            card_id = None
            
        fund_group_id = data.get('fund_group_id')
        if not fund_group_id or str(fund_group_id) == 'None' or str(fund_group_id).strip() == '':
            fund_group_id = None
            
        fund_group_locked = 1 if fund_group_id else 0
        db.execute(
            "UPDATE card_tx SET card_id=?, date=?, name=?, category=?, amount=?, installment=?, memo=?,"
            " category_locked=?, fund_group_id=?, fund_group_locked=? WHERE id=?",
            (card_id, data.get('date'), data.get('name'), category,
             data.get('amount', 0), data.get('installment', 1), data.get('memo'),
             locked, fund_group_id, fund_group_locked, rid)
        )
        db.commit()
        return jsonify({'ok': True})
    db.execute("DELETE FROM card_tx WHERE id = ?", (rid,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/card-tx/bulk', methods=['DELETE'])
def api_card_tx_bulk_delete():
    data = request.json or {}
    ids  = data.get('ids', [])
    if not ids:
        return jsonify({'error': 'ids가 필요합니다'}), 400
    db = get_db()
    placeholders = ','.join('?' * len(ids))
    cur = db.execute(f"DELETE FROM card_tx WHERE id IN ({placeholders})", ids)
    deleted = cur.rowcount
    db.commit(); db.close()
    return jsonify({'ok': True, 'deleted': deleted})


@app.route('/api/card-tx/auto-categorize', methods=['POST'])
def api_card_tx_auto_categorize():
    data    = request.json or {}
    card_id = data.get('card_id')
    year    = data.get('year')
    month   = data.get('month')

    db = get_db()
    query  = "SELECT id, name FROM card_tx WHERE category_locked = 0"
    params = []
    if card_id:
        query += " AND card_id = ?"; params.append(card_id)
    if year and month:
        query += " AND strftime('%Y', date) = ? AND strftime('%m', date) = ?"
        params += [year, month.zfill(2)]

    rows = db.execute(query, params).fetchall()
    updated = 0
    for row in rows:
        hint = _get_category_hint(db, row['name'])
        if hint:
            db.execute("UPDATE card_tx SET category=? WHERE id=?", (hint, row['id']))
            updated += 1
    db.commit()
    db.close()
    return jsonify({'ok': True, 'updated': updated})


# ── API: 주식 ────────────────────────────────────────────────
@app.route('/api/stocks', methods=['GET', 'POST'])
def api_stocks():
    db = get_db()
    if request.method == 'GET':
        rows = db.execute("""
            SELECT s.id, s.name, s.ticker, s.current_price, s.dividend, s.memo,
                   COALESCE(SUM(CASE WHEN t.tx_type='buy'  THEN t.quantity ELSE 0 END), 0) AS buy_qty,
                   COALESCE(SUM(CASE WHEN t.tx_type='sell' THEN t.quantity ELSE 0 END), 0) AS sell_qty,
                   COALESCE(SUM(CASE WHEN t.tx_type='buy'  THEN t.price * t.quantity + t.fee ELSE 0 END), 0) AS total_buy_amount,
                   COALESCE(SUM(CASE WHEN t.tx_type='sell' THEN t.price * t.quantity - t.fee ELSE 0 END), 0) AS total_sell_amount
            FROM stocks s
            LEFT JOIN stock_tx t ON t.stock_id = s.id
            GROUP BY s.id
            ORDER BY s.name
        """).fetchall()
        result = []
        for row in rows:
            r = dict(row)
            qty      = r['buy_qty'] - r['sell_qty']
            avg      = round(r['total_buy_amount'] / r['buy_qty']) if r['buy_qty'] else 0
            eval_amt = round(qty * r['current_price'])
            cost_amt = round(qty * avg)
            r['quantity']       = qty
            r['avg_price']      = avg
            r['eval_amount']    = eval_amt
            r['unrealized_pnl'] = eval_amt - cost_amt
            r['return_rate']    = round((eval_amt - cost_amt) / cost_amt * 100, 2) if cost_amt else 0
            r['realized_pnl']   = round(r['total_sell_amount'] - r['sell_qty'] * avg) if r['sell_qty'] else 0
            result.append(r)
        return jsonify(result)

    data = request.json
    db.execute(
        "INSERT INTO stocks (name, ticker, current_price, dividend, memo) VALUES (?,?,?,?,?)",
        (data.get('name'), data.get('ticker'),
         data.get('current_price', 0), data.get('dividend', 0), data.get('memo'))
    )
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/stocks/<int:rid>', methods=['PUT', 'DELETE'])
def api_stocks_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        db.execute(
            "UPDATE stocks SET name=?, ticker=?, current_price=?, dividend=?, memo=? WHERE id=?",
            (data.get('name'), data.get('ticker'),
             data.get('current_price', 0), data.get('dividend', 0), data.get('memo'), rid)
        )
        db.commit()
        return jsonify({'ok': True})
    db.execute("DELETE FROM stock_tx WHERE stock_id = ?", (rid,))
    db.execute("DELETE FROM stocks WHERE id = ?", (rid,))
    db.commit()
    return jsonify({'ok': True})


# ── API: 주식 거래내역 ────────────────────────────────────────
@app.route('/api/stock-tx', methods=['GET', 'POST'])
def api_stock_tx():
    db = get_db()
    if request.method == 'GET':
        stock_id = request.args.get('stock_id')
        query  = "SELECT t.*, s.name, s.ticker FROM stock_tx t LEFT JOIN stocks s ON t.stock_id = s.id"
        params = []
        if stock_id:
            query += " WHERE t.stock_id = ?"
            params.append(stock_id)
        query += " ORDER BY t.tx_date DESC, t.id DESC"
        rows = db.execute(query, params).fetchall()
        return jsonify(rows_to_list(rows))

    data = request.json
    db.execute(
        "INSERT INTO stock_tx (stock_id, tx_date, tx_type, price, quantity, fee, memo) VALUES (?,?,?,?,?,?,?)",
        (data.get('stock_id'), data.get('tx_date'), data.get('tx_type'),
         data.get('price', 0), data.get('quantity', 0), data.get('fee', 0), data.get('memo'))
    )
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/stock-tx/<int:rid>', methods=['PUT', 'DELETE'])
def api_stock_tx_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        db.execute(
            "UPDATE stock_tx SET stock_id=?, tx_date=?, tx_type=?, price=?, quantity=?, fee=?, memo=? WHERE id=?",
            (data.get('stock_id'), data.get('tx_date'), data.get('tx_type'),
             data.get('price', 0), data.get('quantity', 0), data.get('fee', 0), data.get('memo'), rid)
        )
        db.commit()
        return jsonify({'ok': True})
    db.execute("DELETE FROM stock_tx WHERE id = ?", (rid,))
    db.commit()
    return jsonify({'ok': True})


# ── API: ETF ─────────────────────────────────────────────────
@app.route('/api/etf', methods=['GET', 'POST'])
def api_etf():
    db = get_db()
    if request.method == 'GET':
        rows = db.execute("SELECT * FROM etf ORDER BY name").fetchall()
        return jsonify(rows_to_list(rows))

    data = request.json
    db.execute(
        "INSERT INTO etf (name, ticker, buy_date, buy_price, quantity, current_price, etf_type, memo) VALUES (?,?,?,?,?,?,?,?)",
        (data.get('name'), data.get('ticker'), data.get('buy_date'),
         data.get('buy_price', 0), data.get('quantity', 0),
         data.get('current_price', 0), data.get('etf_type'), data.get('memo'))
    )
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/etf/<int:rid>', methods=['PUT', 'DELETE'])
def api_etf_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        db.execute(
            "UPDATE etf SET name=?, ticker=?, buy_date=?, buy_price=?, quantity=?, current_price=?, etf_type=?, memo=? WHERE id=?",
            (data.get('name'), data.get('ticker'), data.get('buy_date'),
             data.get('buy_price', 0), data.get('quantity', 0),
             data.get('current_price', 0), data.get('etf_type'), data.get('memo'), rid)
        )
        db.commit()
        return jsonify({'ok': True})
    db.execute("DELETE FROM etf WHERE id = ?", (rid,))
    db.commit()
    return jsonify({'ok': True})


# ── API: 코인 ────────────────────────────────────────────────
@app.route('/api/crypto', methods=['GET', 'POST'])
def api_crypto():
    db = get_db()
    if request.method == 'GET':
        rows = db.execute("SELECT * FROM crypto ORDER BY name").fetchall()
        return jsonify(rows_to_list(rows))

    data = request.json
    db.execute(
        "INSERT INTO crypto (name, symbol, exchange, buy_date, buy_price, quantity, current_price, memo) VALUES (?,?,?,?,?,?,?,?)",
        (data.get('name'), data.get('symbol'), data.get('exchange'), data.get('buy_date'),
         data.get('buy_price', 0), data.get('quantity', 0),
         data.get('current_price', 0), data.get('memo'))
    )
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/crypto/<int:rid>', methods=['PUT', 'DELETE'])
def api_crypto_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        db.execute(
            "UPDATE crypto SET name=?, symbol=?, exchange=?, buy_date=?, buy_price=?, quantity=?, current_price=?, memo=? WHERE id=?",
            (data.get('name'), data.get('symbol'), data.get('exchange'), data.get('buy_date'),
             data.get('buy_price', 0), data.get('quantity', 0),
             data.get('current_price', 0), data.get('memo'), rid)
        )
        db.commit()
        return jsonify({'ok': True})
    db.execute("DELETE FROM crypto WHERE id = ?", (rid,))
    db.commit()
    return jsonify({'ok': True})


# ── API: 투자 자산 실시간 가격 동기화 ────────────────────────
@app.route('/api/investments/sync-prices', methods=['POST'])
def api_investments_sync_prices():
    db = get_db()
    results = {'stocks': 0, 'etf': 0, 'crypto': 0, 'errors': []}

    # 1. 주식/ETF 동기화 (yfinance)
    for table in ['stocks', 'etf']:
        items = db.execute(f"SELECT id, ticker FROM {table} WHERE ticker IS NOT NULL AND ticker != ''").fetchall()
        for item in items:
            ticker_raw = item['ticker'].strip()
            ticker_to_search = ticker_raw
            
            # 한국 주식 코드 (6자리 숫자) 처리
            if re.match(r'^\d{6}$', ticker_raw):
                # 우선 KOSPI(.KS) 시도 후 실패 시 KOSDAQ(.KQ) 시도 (yfinance 특징)
                ticker_to_search = ticker_raw + ".KS"
            
            try:
                ticker_obj = yf.Ticker(ticker_to_search)
                # 최근 데이터 가져오기
                hist = ticker_obj.history(period='1d')
                if hist.empty and ticker_to_search.endswith(".KS"):
                    ticker_to_search = ticker_raw + ".KQ"
                    ticker_obj = yf.Ticker(ticker_to_search)
                    hist = ticker_obj.history(period='1d')
                
                if not hist.empty:
                    current_price = int(hist['Close'].iloc[-1])
                    db.execute(f"UPDATE {table} SET current_price=? WHERE id=?", (current_price, item['id']))
                    results[table] += 1
                else:
                    results['errors'].append(f"{ticker_raw}: 시세를 찾을 수 없음")
            except Exception as e:
                results['errors'].append(f"{ticker_raw}: {str(e)}")

    # 2. 코인 동기화 (Upbit)
    crypto_items = db.execute("SELECT id, symbol FROM crypto WHERE symbol IS NOT NULL AND symbol != ''").fetchall()
    for item in crypto_items:
        symbol = item['symbol'].strip().upper()
        try:
            # 업비트 원화 마켓 기준
            res = requests.get(f"https://api.upbit.com/v1/ticker?markets=KRW-{symbol}", timeout=5)
            if res.status_code == 200:
                data = res.json()
                if data:
                    current_price = float(data[0]['trade_price'])
                    db.execute("UPDATE crypto SET current_price=? WHERE id=?", (current_price, item['id']))
                    results['crypto'] += 1
                else:
                    results['errors'].append(f"{symbol}: 업비트 원화 마켓에 없음")
            else:
                results['errors'].append(f"{symbol}: API 호출 실패")
        except Exception as e:
            results['errors'].append(f"{symbol}: {str(e)}")

    db.commit()
    db.close()
    return jsonify(results)


# ── API: 거주지 ──────────────────────────────────────────────
@app.route('/api/residence', methods=['GET', 'POST'])
def api_residence():
    db = get_db()
    if request.method == 'GET':
        rows = db.execute("SELECT * FROM residence ORDER BY id DESC").fetchall()
        return jsonify(rows_to_list(rows))

    data = request.json
    db.execute(
        "INSERT INTO residence (address, deposit, monthly_rent, maintenance, start_date, end_date) VALUES (?,?,?,?,?,?)",
        (data.get('address'), data.get('deposit', 0), data.get('monthly_rent', 0),
         data.get('maintenance', 0), data.get('start_date'), data.get('end_date'))
    )
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/residence/<int:rid>', methods=['PUT', 'DELETE'])
def api_residence_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        db.execute(
            "UPDATE residence SET address=?, deposit=?, monthly_rent=?, maintenance=?, start_date=?, end_date=? WHERE id=?",
            (data.get('address'), data.get('deposit', 0), data.get('monthly_rent', 0),
             data.get('maintenance', 0), data.get('start_date'), data.get('end_date'), rid)
        )
        db.commit()
        return jsonify({'ok': True})
    db.execute("DELETE FROM residence WHERE id = ?", (rid,))
    db.commit()
    return jsonify({'ok': True})


# ── API: 부동산 ──────────────────────────────────────────────
def _re_enrich(db, rows):
    """real_estate 목록에 현재 계약 정보·비용 집계·실수익률 추가"""
    result = []
    for r in rows:
        rid = r['id']
        # 현재 유효 계약 (end_date 가장 최근)
        contract = db.execute(
            "SELECT * FROM tenant_contracts WHERE real_estate_id=? ORDER BY end_date DESC LIMIT 1", (rid,)
        ).fetchone()
        # 취득비용 합계
        acq_cost = db.execute(
            "SELECT COALESCE(SUM(amount),0) as v FROM property_costs "
            "WHERE real_estate_id=? AND cost_type='취득비용'", (rid,)
        ).fetchone()['v']
        # 순손익에 반영될 비용/수익 합계 (amount: 수익=양수, 비용=음수)
        net_extra = db.execute(
            "SELECT COALESCE(SUM(CASE WHEN cost_type='임대수익' THEN amount ELSE -amount END),0) as v "
            "FROM property_costs WHERE real_estate_id=?", (rid,)
        ).fetchone()['v']

        deposit = contract['deposit'] if contract else 0
        purchase = r['purchase_price']
        current  = r['current_price']
        real_inv = purchase - deposit + acq_cost   # 실투자금
        net_gain = (current - purchase) + net_extra  # 순손익
        real_roi = round(net_gain / real_inv * 100, 1) if real_inv > 0 else None

        row = dict(r)
        row['contract_type']  = contract['contract_type'] if contract else None
        row['deposit']        = deposit
        row['monthly_rent']   = contract['monthly_rent'] if contract else 0
        row['contract_end']   = contract['end_date'] if contract else None
        row['real_inv']       = real_inv
        row['net_gain']       = net_gain
        row['real_roi']       = real_roi
        result.append(row)
    return result


@app.route('/api/real-estate', methods=['GET', 'POST'])
def api_real_estate():
    db = get_db()
    if request.method == 'GET':
        rows = db.execute("SELECT * FROM real_estate ORDER BY name").fetchall()
        result = _re_enrich(db, rows)
        return jsonify(result)

    data = request.json
    db.execute(
        "INSERT INTO real_estate (name, re_type, purchase_date, purchase_price, current_price, memo) VALUES (?,?,?,?,?,?)",
        (data.get('name'), data.get('re_type'), data.get('purchase_date'),
         data.get('purchase_price', 0), data.get('current_price', 0), data.get('memo'))
    )
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/real-estate/<int:rid>', methods=['PUT', 'DELETE'])
def api_real_estate_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        db.execute(
            "UPDATE real_estate SET name=?, re_type=?, purchase_date=?, purchase_price=?, current_price=?, memo=? WHERE id=?",
            (data.get('name'), data.get('re_type'), data.get('purchase_date'),
             data.get('purchase_price', 0), data.get('current_price', 0), data.get('memo'), rid)
        )
        db.commit()
        return jsonify({'ok': True})
    db.execute("DELETE FROM tenant_contracts WHERE real_estate_id = ?", (rid,))
    db.execute("DELETE FROM property_costs WHERE real_estate_id = ?", (rid,))
    db.execute("DELETE FROM real_estate WHERE id = ?", (rid,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/re-summary')
def api_re_summary():
    db = get_db()
    rows = db.execute("SELECT * FROM real_estate").fetchall()
    enriched = _re_enrich(db, rows)
    db.close()
    total_purchase = sum(r['purchase_price'] for r in enriched)
    total_deposit  = sum(r['deposit'] for r in enriched)
    total_real_inv = sum(r['real_inv'] for r in enriched)
    total_net_gain = sum(r['net_gain'] for r in enriched)
    avg_roi = round(total_net_gain / total_real_inv * 100, 1) if total_real_inv > 0 else None
    return jsonify({
        'count': len(enriched),
        'total_purchase': total_purchase,
        'total_deposit': total_deposit,
        'total_real_inv': total_real_inv,
        'total_net_gain': total_net_gain,
        'avg_roi': avg_roi,
    })


@app.route('/api/re-expiring')
def api_re_expiring():
    """3개월 이내 만료 계약 목록"""
    db = get_db()
    rows = db.execute("""
        SELECT c.*, r.name as re_name
        FROM tenant_contracts c
        JOIN real_estate r ON c.real_estate_id = r.id
        WHERE c.end_date IS NOT NULL
          AND c.end_date >= date('now')
          AND c.end_date <= date('now', '+3 months')
        ORDER BY c.end_date
    """).fetchall()
    db.close()
    return jsonify(rows_to_list(rows))


@app.route('/api/tenant-contracts', methods=['GET', 'POST'])
def api_tenant_contracts():
    db = get_db()
    if request.method == 'GET':
        rid = request.args.get('real_estate_id')
        rows = db.execute(
            "SELECT * FROM tenant_contracts WHERE real_estate_id=? ORDER BY end_date DESC", (rid,)
        ).fetchall()
        return jsonify(rows_to_list(rows))
    d = request.json or {}
    db.execute(
        "INSERT INTO tenant_contracts (real_estate_id, contract_type, deposit, monthly_rent, start_date, end_date, memo)"
        " VALUES (?,?,?,?,?,?,?)",
        (d.get('real_estate_id'), d.get('contract_type'), d.get('deposit', 0),
         d.get('monthly_rent', 0), d.get('start_date'), d.get('end_date'), d.get('memo'))
    )
    db.commit(); db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/tenant-contracts/<int:rid>', methods=['PUT', 'DELETE'])
def api_tenant_contracts_detail(rid):
    db = get_db()
    if request.method == 'DELETE':
        db.execute("DELETE FROM tenant_contracts WHERE id=?", (rid,))
        db.commit(); db.close()
        return jsonify({'ok': True})
    d = request.json or {}
    db.execute(
        "UPDATE tenant_contracts SET contract_type=?, deposit=?, monthly_rent=?, start_date=?, end_date=?, memo=? WHERE id=?",
        (d.get('contract_type'), d.get('deposit', 0), d.get('monthly_rent', 0),
         d.get('start_date'), d.get('end_date'), d.get('memo'), rid)
    )
    db.commit(); db.close()
    return jsonify({'ok': True})


@app.route('/api/property-costs', methods=['GET', 'POST'])
def api_property_costs():
    db = get_db()
    if request.method == 'GET':
        rid = request.args.get('real_estate_id')
        rows = db.execute(
            "SELECT * FROM property_costs WHERE real_estate_id=? ORDER BY date DESC", (rid,)
        ).fetchall()
        return jsonify(rows_to_list(rows))
    d = request.json or {}
    db.execute(
        "INSERT INTO property_costs (real_estate_id, cost_type, name, amount, date, memo) VALUES (?,?,?,?,?,?)",
        (d.get('real_estate_id'), d.get('cost_type'), d.get('name'),
         d.get('amount', 0), d.get('date'), d.get('memo'))
    )
    db.commit(); db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/property-costs/<int:rid>', methods=['PUT', 'DELETE'])
def api_property_costs_detail(rid):
    db = get_db()
    if request.method == 'DELETE':
        db.execute("DELETE FROM property_costs WHERE id=?", (rid,))
        db.commit(); db.close()
        return jsonify({'ok': True})
    d = request.json or {}
    db.execute(
        "UPDATE property_costs SET cost_type=?, name=?, amount=?, date=?, memo=? WHERE id=?",
        (d.get('cost_type'), d.get('name'), d.get('amount', 0),
         d.get('date'), d.get('memo'), rid)
    )
    db.commit(); db.close()
    return jsonify({'ok': True})


# ── API: 대출 ────────────────────────────────────────────────
@app.route('/api/loans', methods=['GET', 'POST'])
def api_loans():
    db = get_db()
    if request.method == 'GET':
        rows = db.execute("SELECT * FROM loans ORDER BY name").fetchall()
        return jsonify(rows_to_list(rows))

    data = request.json
    db.execute(
        "INSERT INTO loans (name, institution, principal, remaining, monthly_payment, interest_rate, loan_date, end_date, memo) VALUES (?,?,?,?,?,?,?,?,?)",
        (data.get('name'), data.get('institution'), data.get('principal', 0),
         data.get('remaining', 0), data.get('monthly_payment', 0),
         data.get('interest_rate', 0), data.get('loan_date'), data.get('end_date'), data.get('memo'))
    )
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/loans/<int:rid>', methods=['PUT', 'DELETE'])
def api_loans_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        db.execute(
            "UPDATE loans SET name=?, institution=?, principal=?, remaining=?, monthly_payment=?, interest_rate=?, loan_date=?, end_date=?, memo=? WHERE id=?",
            (data.get('name'), data.get('institution'), data.get('principal', 0),
             data.get('remaining', 0), data.get('monthly_payment', 0),
             data.get('interest_rate', 0), data.get('loan_date'), data.get('end_date'), data.get('memo'), rid)
        )
        db.commit()
        return jsonify({'ok': True})
    db.execute("DELETE FROM loans WHERE id = ?", (rid,))
    db.commit()
    return jsonify({'ok': True})


# ── API: 연금 ────────────────────────────────────────────────
@app.route('/api/pension', methods=['GET', 'POST'])
def api_pension():
    db = get_db()
    if request.method == 'GET':
        rows = db.execute("SELECT * FROM pension ORDER BY pension_type, name").fetchall()
        return jsonify(rows_to_list(rows))

    data = request.json
    db.execute(
        "INSERT INTO pension (pension_type, name, institution, monthly_payment, accumulated, return_rate, memo) VALUES (?,?,?,?,?,?,?)",
        (data.get('pension_type'), data.get('name'), data.get('institution'),
         data.get('monthly_payment', 0), data.get('accumulated', 0),
         data.get('return_rate', 0), data.get('memo'))
    )
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/pension/<int:rid>', methods=['PUT', 'DELETE'])
def api_pension_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        db.execute(
            "UPDATE pension SET pension_type=?, name=?, institution=?, monthly_payment=?, accumulated=?, return_rate=?, memo=? WHERE id=?",
            (data.get('pension_type'), data.get('name'), data.get('institution'),
             data.get('monthly_payment', 0), data.get('accumulated', 0),
             data.get('return_rate', 0), data.get('memo'), rid)
        )
        db.commit()
        return jsonify({'ok': True})
    db.execute("DELETE FROM pension WHERE id = ?", (rid,))
    db.commit()
    return jsonify({'ok': True})


# ── API: 목표저축 ────────────────────────────────────────────
@app.route('/api/goals', methods=['GET', 'POST'])
def api_goals():
    db = get_db()
    if request.method == 'GET':
        rows = db.execute("SELECT * FROM goals ORDER BY target_date").fetchall()
        return jsonify(rows_to_list(rows))

    data = request.json
    db.execute(
        "INSERT INTO goals (name, target_amount, current_amount, monthly_saving, target_date, memo) VALUES (?,?,?,?,?,?)",
        (data.get('name'), data.get('target_amount', 0), data.get('current_amount', 0),
         data.get('monthly_saving', 0), data.get('target_date'), data.get('memo'))
    )
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/goals/<int:rid>', methods=['PUT', 'DELETE'])
def api_goals_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        db.execute(
            "UPDATE goals SET name=?, target_amount=?, current_amount=?, monthly_saving=?, target_date=?, memo=? WHERE id=?",
            (data.get('name'), data.get('target_amount', 0), data.get('current_amount', 0),
             data.get('monthly_saving', 0), data.get('target_date'), data.get('memo'), rid)
        )
        db.commit()
        return jsonify({'ok': True})
    db.execute("DELETE FROM goals WHERE id = ?", (rid,))
    db.commit()
    return jsonify({'ok': True})


# ── API: 현금/예금 ───────────────────────────────────────────
@app.route('/api/cash-deposits', methods=['GET', 'POST'])
def api_cash_deposits():
    db = get_db()
    if request.method == 'GET':
        rows = db.execute("SELECT * FROM cash_deposits ORDER BY name").fetchall()
        return jsonify(rows_to_list(rows))

    data = request.json
    today = date.today().isoformat()
    db.execute(
        "INSERT INTO cash_deposits (name, amount, memo, updated_date) VALUES (?,?,?,?)",
        (data.get('name'), data.get('amount', 0), data.get('memo'), today)
    )
    db.commit()
    db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/cash-deposits/<int:rid>', methods=['PUT', 'DELETE'])
def api_cash_deposits_detail(rid):
    db = get_db()
    if request.method == 'PUT':
        data = request.json
        today = date.today().isoformat()
        db.execute(
            "UPDATE cash_deposits SET name=?, amount=?, memo=?, updated_date=? WHERE id=?",
            (data.get('name'), data.get('amount', 0), data.get('memo'), today, rid)
        )
        db.commit()
        return jsonify({'ok': True})
    db.execute("DELETE FROM cash_deposits WHERE id = ?", (rid,))
    db.commit()
    return jsonify({'ok': True})


# ── API: 대시보드 집계 ───────────────────────────────────────
@app.route('/api/dashboard')
def api_dashboard():
    db = get_db()
    today = date.today()
    year  = request.args.get('year',  today.strftime('%Y'))
    month = request.args.get('month', today.strftime('%m'))
    ym    = f"{year}-{month.zfill(2)}"

    # 이번달 수입 합계
    income_total = db.execute(
        "SELECT COALESCE(SUM(amount),0) as total FROM income WHERE strftime('%Y-%m', date) = ?", (ym,)
    ).fetchone()['total']

    # 이번달 지출 합계
    expense_total = db.execute(
        "SELECT COALESCE(SUM(amount),0) as total FROM budget WHERE strftime('%Y-%m', date) = ?", (ym,)
    ).fetchone()['total']

    # 이번달 카드 지출
    card_total = db.execute(
        "SELECT COALESCE(SUM(amount),0) as total FROM card_tx WHERE strftime('%Y-%m', date) = ?", (ym,)
    ).fetchone()['total']

    # 주식 평가액 (stock_tx 기반)
    stocks_val = db.execute("""
        SELECT COALESCE(SUM(s.current_price * (
            SELECT COALESCE(SUM(CASE WHEN tx_type='buy' THEN quantity ELSE -quantity END), 0)
            FROM stock_tx WHERE stock_id = s.id
        )), 0) AS val FROM stocks s
    """).fetchone()['val']

    # ETF 평가액
    etf_val = db.execute(
        "SELECT COALESCE(SUM(current_price * quantity),0) as val FROM etf"
    ).fetchone()['val']

    # 코인 평가액
    crypto_val = db.execute(
        "SELECT COALESCE(SUM(current_price * quantity),0) as val FROM crypto"
    ).fetchone()['val']

    # 부동산 현재가
    re_val = db.execute(
        "SELECT COALESCE(SUM(current_price),0) as val FROM real_estate"
    ).fetchone()['val']

    # 연금 누적액
    pension_val = db.execute(
        "SELECT COALESCE(SUM(accumulated),0) as val FROM pension"
    ).fetchone()['val']

    # 현금/예금
    cash_val = db.execute(
        "SELECT COALESCE(SUM(amount),0) as val FROM cash_deposits"
    ).fetchone()['val']

    # 대출 잔액
    loan_total = db.execute(
        "SELECT COALESCE(SUM(remaining),0) as total FROM loans"
    ).fetchone()['total']

    total_assets = stocks_val + etf_val + crypto_val + re_val + pension_val + cash_val
    net_worth = total_assets - loan_total

    # 이번달 수입 카테고리별
    income_by_cat = db.execute(
        "SELECT category, SUM(amount) as total FROM income WHERE strftime('%Y-%m', date) = ? GROUP BY category",
        (ym,)
    ).fetchall()

    # 이번달 지출 카테고리별 (수동 지출 + 카드 지출 합산)
    expense_by_cat = db.execute("""
        SELECT category, SUM(amount) as total FROM (
            SELECT COALESCE(category, '미분류') as category, amount FROM budget WHERE strftime('%Y-%m', date) = ?
            UNION ALL
            SELECT COALESCE(category, '미분류') as category, amount FROM card_tx WHERE strftime('%Y-%m', date) = ?
        ) GROUP BY category ORDER BY total DESC
    """, (ym, ym)).fetchall()

    # 대출 목록
    loans_list = db.execute(
        "SELECT name, remaining FROM loans ORDER BY remaining DESC"
    ).fetchall()

    # 목표저축 목록
    goals_list = db.execute(
        "SELECT name, target_amount, current_amount FROM goals ORDER BY target_date"
    ).fetchall()

    # 투자 수익률
    stocks_cost = db.execute(
        "SELECT COALESCE(SUM(price * quantity + fee),0) AS c FROM stock_tx WHERE tx_type='buy'"
    ).fetchone()['c']
    etf_cost    = db.execute("SELECT COALESCE(SUM(buy_price * quantity),0) as c FROM etf").fetchone()['c']
    crypto_cost = db.execute("SELECT COALESCE(SUM(buy_price * quantity),0) as c FROM crypto").fetchone()['c']

    db.close()

    return jsonify({
        'income_total':    income_total,
        'expense_total':   expense_total + card_total,
        'net_worth':       net_worth,
        'total_assets':    total_assets,
        'loan_total':      loan_total,
        'asset_breakdown': {
            'stocks':  stocks_val,
            'etf':     etf_val,
            'crypto':  crypto_val,
            'realestate': re_val,
            'pension': pension_val,
            'cash':    cash_val,
        },
        'income_by_cat':  rows_to_list(income_by_cat),
        'expense_by_cat': rows_to_list(expense_by_cat),
        'loans':          rows_to_list(loans_list),
        'goals':          rows_to_list(goals_list),
        'investment_returns': {
            'stocks':  {'cost': stocks_cost, 'value': stocks_val},
            'etf':     {'cost': etf_cost,    'value': etf_val},
            'crypto':  {'cost': crypto_cost, 'value': crypto_val},
        },
    })


# ── API: 월별 결산 ───────────────────────────────────────────
@app.route('/api/monthly-summary')
def api_monthly_summary():
    db = get_db()
    year = request.args.get('year', date.today().strftime('%Y'))

    months = []
    for m in range(1, 13):
        ym = f"{year}-{m:02d}"
        inc = db.execute(
            "SELECT COALESCE(SUM(amount),0) as t FROM income WHERE strftime('%Y-%m', date) = ?", (ym,)
        ).fetchone()['t']
        exp = db.execute(
            "SELECT COALESCE(SUM(amount),0) as t FROM budget WHERE strftime('%Y-%m', date) = ?", (ym,)
        ).fetchone()['t']
        card = db.execute(
            "SELECT COALESCE(SUM(amount),0) as t FROM card_tx WHERE strftime('%Y-%m', date) = ?", (ym,)
        ).fetchone()['t']
        months.append({
            'month':   ym,
            'income':  inc,
            'expense': exp + card,
            'saving':  inc - (exp + card),
        })

    db.close()
    return jsonify(months)


# ── 카드 엑셀 가져오기 ───────────────────────────────────────
_HINTS = {
    'date':        ['이용일', '이용일시', '거래일', '거래일자', '이용일자', '날짜', '결제일', '승인일', '사용일', '거래 일시'],
    'name':        ['이용가맹점', '가맹점명', '상호명', '가맹점', '이용처', '거래처', '내용', '적요', '사용처', '거래내용'],
    'amount':      ['이용금액', '결제금액', '거래금액', '금액', '승인금액', '사용금액', '국내이용금액'],
    'installment': ['할부', '할부개월', '할부개월수', '분할', '할부기간'],
    'category':    ['업종', '카테고리', '이용구분', '업종명', '분류'],
}

def _detect_header(rows):
    all_hints = [h for hs in _HINTS.values() for h in hs]
    best, idx = 0, 0
    for i, row in enumerate(rows[:15]):
        score = sum(1 for c in row if any(h in str(c) for h in all_hints))
        if score > best:
            best, idx = score, i
    return idx

def _detect_mapping(headers):
    m = {}
    for field, hints in _HINTS.items():
        for h in headers:
            if any(hint in str(h) for hint in hints):
                m[field] = h; break
    return m

def _parse_date(val):
    if val is None: return None
    if hasattr(val, 'strftime'): return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    
    # 2026년 03월 01일 (공백 포함 가능하므로 먼저 처리)
    m = re.match(r'(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일', s)
    if m: return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    
    # ISO 형식 (T 또는 공백으로 시간 분리) 처리
    s_date = s.split(' ')[0].split('T')[0]
    
    # YYYY-MM-DD, YYYY.MM.DD, YYYY/MM/DD
    m = re.match(r'(\d{4})[-./](\d{1,2})[-./](\d{1,2})', s_date)
    if m: return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    
    # YYYYMMDD
    if len(s_date) == 8 and s_date.isdigit():
        return f"{s_date[:4]}-{s_date[4:6]}-{s_date[6:]}"
        
    return None

def _parse_amount(val):
    if val is None: return 0
    if isinstance(val, (int, float)): return int(abs(val))
    try: return int(abs(float(str(val).replace(',', '').replace(' ', ''))))
    except: return 0

def _parse_file(file_bytes, filename):
    rows = []
    name = filename.lower()
    error_msg = None

    try:
        if name.endswith('.csv'):
            success = False
            for enc in ['utf-8-sig', 'cp949', 'euc-kr', 'utf-8']:
                try:
                    text = file_bytes.decode(enc)
                    rows = [[str(c).strip() for c in r]
                            for r in csv.reader(io.StringIO(text)) if any(c for c in r)]
                    if rows:
                        success = True
                        break
                except: continue
            if not success:
                error_msg = "CSV 파일 인코딩을 인식할 수 없습니다. (UTF-8 또는 CP949 형식이 필요합니다)"
        elif name.endswith('.xls'):
            if not HAS_XLRD:
                error_msg = "구형 엑셀(.xls) 처리를 위한 라이브러리(xlrd)가 설치되어 있지 않습니다."
            else:
                try:
                    wb = xlrd.open_workbook(file_contents=file_bytes)
                    ws = wb.sheet_by_index(0)
                    for i in range(ws.nrows):
                        row_vals = []
                        for j in range(ws.ncols):
                            cell = ws.cell(i, j)
                            if cell.ctype == xlrd.XL_CELL_DATE:
                                try:
                                    t = xlrd.xldate.xldate_as_tuple(cell.value, wb.datemode)
                                    row_vals.append(datetime(*t[:6]))
                                except: row_vals.append(cell.value)
                            else: row_vals.append(cell.value)
                        if any(v is not None and str(v).strip() for v in row_vals):
                            rows.append(row_vals)
                except Exception as e:
                    # HTML 형식의 가짜 엑셀 처리
                    if b'<html' in file_bytes[:1024].lower() or b'<table' in file_bytes[:1024].lower():
                        text = None
                        for enc in ['utf-8-sig', 'cp949', 'euc-kr', 'utf-8']:
                            try:
                                text = file_bytes.decode(enc)
                                break
                            except: continue
                        if text:
                            tr_matches = re.findall(r'<tr[^>]*>(.*?)</tr>', text, re.I | re.S)
                            for tr in tr_matches:
                                td_matches = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', tr, re.I | re.S)
                                row_vals = [re.sub(r'<[^>]+>', '', td).strip() for td in td_matches]
                                if any(row_vals): rows.append(row_vals)
                    if not rows:
                        error_msg = f"Excel(.xls) 파싱 실패: {str(e)}"
        elif name.endswith('.xlsx'):
            if not HAS_OPENPYXL:
                error_msg = "엑셀(.xlsx) 처리를 위한 라이브러리(openpyxl)가 설치되어 있지 않습니다."
            else:
                try:
                    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                    ws = wb.active
                    rows = [[cell.value for cell in row] for row in ws.iter_rows()
                            if any(c.value is not None for c in row)]
                except Exception as e:
                    error_msg = f"Excel(.xlsx) 파일을 읽는 중 오류가 발생했습니다: {str(e)}"
        else:
            error_msg = "지원하지 않는 파일 형식입니다. (.xlsx, .xls, .csv 파일을 올려주세요)"

        if not error_msg and not rows:
            error_msg = "파일에서 읽을 수 있는 데이터가 없습니다."

        if error_msg:
            return None, None, error_msg

        hi = _detect_header(rows)
        headers = [str(h).strip() if h is not None else f'컬럼{i}' for i, h in enumerate(rows[hi])]
        data    = [list(row) for row in rows[hi+1:]
                   if any(v is not None and str(v).strip() for v in row)]
        return headers, data, None

    except Exception as e:
        return None, None, f"파일 처리 중 예상치 못한 오류 발생: {str(e)}"


@app.route('/api/card-excel/preview', methods=['POST'])
def api_card_excel_preview():
    f = request.files.get('file')
    if not f: return jsonify({'error': '파일이 없습니다'}), 400
    headers, data, error = _parse_file(f.read(), f.filename)
    if error: return jsonify({'error': error}), 400
    sample = [
        {h: (str(row[i]).strip() if i < len(row) and row[i] is not None else '')
         for i, h in enumerate(headers)}
        for row in data[:5]
    ]
    all_rows = [
        [(str(row[i]).strip() if i < len(row) and row[i] is not None else '') if not hasattr(row[i] if i < len(row) else None, 'strftime')
         else (row[i].strftime('%Y-%m-%d') if i < len(row) and row[i] is not None else '')
         for i in range(len(headers))]
        for row in data
    ]
    return jsonify({'headers': headers, 'mapping': _detect_mapping(headers),
                    'sample': sample, 'total': len(data), 'all_rows': all_rows})


@app.route('/api/card-excel/mapping/<int:card_id>', methods=['GET', 'POST'])
def api_card_excel_mapping(card_id):
    db = get_db()
    if request.method == 'GET':
        row = db.execute("SELECT mapping FROM card_mappings WHERE card_id=?", (card_id,)).fetchone()
        return jsonify(json.loads(row['mapping']) if row else {})
    db.execute(
        """
        INSERT INTO card_mappings (card_id, mapping)
        VALUES (?, ?)
        ON CONFLICT (card_id) DO UPDATE SET mapping = EXCLUDED.mapping
        """,
        (card_id, json.dumps(request.json or {}))
    )
    db.commit(); db.close()
    return jsonify({'ok': True})


@app.route('/api/card-excel/import', methods=['POST'])
def api_card_excel_import():
    data     = request.json or {}
    card_id  = data.get('card_id')
    mapping  = data.get('mapping', {})
    headers  = data.get('headers', [])
    all_rows = data.get('all_rows', [])
    if not card_id or not mapping.get('date') or not mapping.get('amount'):
        return jsonify({'error': '카드, 날짜, 금액 컬럼은 필수입니다'}), 400

    hi = {h: i for i, h in enumerate(headers)}
    def get(row, col): return row[hi[col]] if col and col in hi and hi[col] < len(row) else ''

    db = get_db()
    inserted = duplicate = 0
    skipped_details = []
    
    for i, row in enumerate(all_rows):
        raw_date = get(row, mapping['date'])
        raw_amount = get(row, mapping['amount'])
        
        date_str = _parse_date(raw_date)
        amount   = _parse_amount(raw_amount)
        name     = str(get(row, mapping.get('name', '')) or '').strip()
        inst     = _parse_amount(get(row, mapping.get('installment', ''))) or 1
        category = str(get(row, mapping.get('category', '')) or '').strip()
        
        # 스킵 사유 확인
        if not date_str:
            skipped_details.append({'row': i+1, 'reason': '날짜 형식 오류', 'data': raw_date})
            continue
        if amount <= 0:
            skipped_details.append({'row': i+1, 'reason': '금액 누락(0 이하)', 'data': raw_amount})
            continue
            
        if db.execute("SELECT id FROM card_tx WHERE card_id=? AND date=? AND name=? AND amount=?",
                      (card_id, date_str, name, amount)).fetchone():
            duplicate += 1
            continue
            
        # 카테고리가 없으면 힌트 자동 적용
        if not category and name:
            category = _get_category_hint(db, name)
        db.execute("INSERT INTO card_tx (card_id,date,name,category,amount,installment) VALUES (?,?,?,?,?,?)",
                   (card_id, date_str, name, category, amount, inst))
        inserted += 1
        
    db.commit()
    return jsonify({
        'ok': True, 
        'inserted': inserted, 
        'duplicate': duplicate, 
        'skipped': len(skipped_details),
        'skipped_details': skipped_details
    })


# ── API: 카테고리 자동 힌트 ──────────────────────────────────
def _get_category_hint(db, name):
    """가맹점명으로 카테고리 추천 (히스토리 → 키워드 규칙 순)"""
    name = (name or '').strip()
    if not name:
        return ''
    # 1. 히스토리: 동일 가맹점의 가장 최근 카테고리
    row = db.execute(
        "SELECT category FROM card_tx WHERE name=? AND category IS NOT NULL AND category!='' "
        "ORDER BY date DESC LIMIT 1", (name,)
    ).fetchone()
    if row:
        return row['category']
    # 2. 키워드 규칙 (긴 키워드 우선)
    rules = db.execute(
        "SELECT keyword, category FROM card_category_rules ORDER BY LENGTH(keyword) DESC"
    ).fetchall()
    name_lower = name.lower()
    for rule in rules:
        if rule['keyword'].lower() in name_lower:
            return rule['category']
    return ''


@app.route('/api/card-category-hint')
def api_card_category_hint():
    name = request.args.get('name', '').strip()
    if not name:
        return jsonify({'category': ''})
    db = get_db()
    category = _get_category_hint(db, name)
    db.close()
    return jsonify({'category': category})


@app.route('/api/card-category-rules', methods=['GET', 'POST'])
def api_card_category_rules():
    db = get_db()
    if request.method == 'GET':
        rows = db.execute("SELECT * FROM card_category_rules ORDER BY keyword").fetchall()
        return jsonify(rows_to_list(rows))
    d = request.json or {}
    keyword  = d.get('keyword', '').strip()
    category = d.get('category', '').strip()
    force    = d.get('force', False)
    
    if not keyword:
        return jsonify({'error': '키워드가 필요합니다.'}), 400

    existing = db.execute("SELECT id FROM card_category_rules WHERE keyword=?", (keyword,)).fetchone()
    if existing and not force:
        return jsonify({'error': 'duplicate', 'message': f"'{keyword}' 규칙이 이미 존재합니다. 덮어쓸까요?"}), 409
    
    if existing:
        db.execute("UPDATE card_category_rules SET category=? WHERE id=?", (category, existing['id']))
    else:
        db.execute("INSERT INTO card_category_rules (keyword, category) VALUES (?, ?)", (keyword, category))
    
    db.commit(); db.close()
    return jsonify({'ok': True})


@app.route('/api/categories', methods=['GET', 'POST'])
def api_categories():
    db = get_db()
    if request.method == 'GET':
        rows = db.execute("SELECT * FROM categories ORDER BY sort_order, name").fetchall()
        return jsonify(rows_to_list(rows))
    d = request.json or {}
    max_order = db.execute("SELECT COALESCE(MAX(sort_order), -1) AS max_order FROM categories").fetchone()['max_order']
    db.execute(
        "INSERT INTO categories (name, sort_order) VALUES (?, ?) ON CONFLICT (name) DO NOTHING",
        (d.get('name', '').strip(), max_order + 1)
    )
    db.commit(); db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/categories/<int:rid>', methods=['PUT', 'DELETE'])
def api_categories_detail(rid):
    db = get_db()
    if request.method == 'DELETE':
        db.execute("DELETE FROM categories WHERE id=?", (rid,))
        db.commit(); db.close()
        return jsonify({'ok': True})
    d = request.json or {}
    db.execute("UPDATE categories SET name=? WHERE id=?", (d.get('name', '').strip(), rid))
    db.commit(); db.close()
    return jsonify({'ok': True})


@app.route('/api/card-category-rules/<int:rid>', methods=['PUT', 'DELETE'])
def api_card_category_rules_detail(rid):
    db = get_db()
    if request.method == 'DELETE':
        db.execute("DELETE FROM card_category_rules WHERE id=?", (rid,))
        db.commit(); db.close()
        return jsonify({'ok': True})
    d = request.json or {}
    db.execute("UPDATE card_category_rules SET keyword=?, category=? WHERE id=?",
               (d.get('keyword', ''), d.get('category', ''), rid))
    db.commit(); db.close()
    return jsonify({'ok': True})


# ── API: 자금 그룹 ────────────────────────────────────────────
def _get_fund_group_hint(db, name):
    """가맹점명으로 자금 그룹 추천 (히스토리 → 키워드 규칙 순)"""
    name = (name or '').strip()
    if not name:
        return None
    # 1. 히스토리: 동일 가맹점의 가장 최근 자금 그룹
    row = db.execute(
        "SELECT fund_group_id FROM card_tx WHERE name=? AND fund_group_id IS NOT NULL "
        "ORDER BY date DESC LIMIT 1", (name,)
    ).fetchone()
    if row:
        return row['fund_group_id']
    # 2. 키워드 규칙 (긴 키워드 우선)
    rules = db.execute(
        "SELECT keyword, fund_group_id FROM fund_group_rules ORDER BY LENGTH(keyword) DESC"
    ).fetchall()
    name_lower = name.lower()
    for rule in rules:
        if rule['keyword'].lower() in name_lower:
            return rule['fund_group_id']
    return None


@app.route('/fund-management')
def fund_management():
    return render_template('fund_management.html')


@app.route('/api/fund-groups', methods=['GET', 'POST'])
def api_fund_groups():
    db = get_db()
    if request.method == 'GET':
        rows = db.execute("SELECT * FROM fund_groups ORDER BY sort_order, name").fetchall()
        return jsonify(rows_to_list(rows))
    d = request.json or {}
    max_order = db.execute("SELECT COALESCE(MAX(sort_order), -1) AS max_order FROM fund_groups").fetchone()['max_order']
    db.execute(
        "INSERT INTO fund_groups (name, sort_order) VALUES (?, ?) ON CONFLICT (name) DO NOTHING",
        (d.get('name', '').strip(), max_order + 1)
    )
    db.commit(); db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/fund-groups/<int:rid>', methods=['PUT', 'DELETE'])
def api_fund_groups_detail(rid):
    db = get_db()
    if request.method == 'DELETE':
        db.execute("DELETE FROM fund_group_rules WHERE fund_group_id = ?", (rid,))
        db.execute("DELETE FROM monthly_fund_budgets WHERE fund_group_id = ?", (rid,))
        db.execute("UPDATE card_tx SET fund_group_id = NULL WHERE fund_group_id = ?", (rid,))
        db.execute("DELETE FROM fund_groups WHERE id=?", (rid,))
        db.commit()
        return jsonify({'ok': True})
    d = request.json or {}
    db.execute("UPDATE fund_groups SET name=? WHERE id=?", (d.get('name', '').strip(), rid))
    db.commit(); db.close()
    return jsonify({'ok': True})


@app.route('/api/fund-group-rules', methods=['GET', 'POST'])
def api_fund_group_rules():
    db = get_db()
    if request.method == 'GET':
        rows = db.execute(
            "SELECT r.*, g.name as fund_group_name FROM fund_group_rules r "
            "LEFT JOIN fund_groups g ON r.fund_group_id = g.id ORDER BY r.keyword"
        ).fetchall()
        return jsonify(rows_to_list(rows))
    d = request.json or {}
    keyword = d.get('keyword', '').strip()
    f_gid   = d.get('fund_group_id')
    force   = d.get('force', False)
    
    if not keyword:
        return jsonify({'error': '키워드가 필요합니다.'}), 400

    existing = db.execute("SELECT id FROM fund_group_rules WHERE keyword=?", (keyword,)).fetchone()
    if existing and not force:
        return jsonify({'error': 'duplicate', 'message': f"'{keyword}' 규칙이 이미 존재합니다. 덮어쓸까요?"}), 409

    if existing:
        db.execute("UPDATE fund_group_rules SET fund_group_id=? WHERE id=?", (f_gid, existing['id']))
    else:
        db.execute("INSERT INTO fund_group_rules (keyword, fund_group_id) VALUES (?, ?)", (keyword, f_gid))
    
    db.commit(); db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/fund-group-rules/<int:rid>', methods=['PUT', 'DELETE'])
def api_fund_group_rules_detail(rid):
    db = get_db()
    if request.method == 'DELETE':
        db.execute("DELETE FROM fund_group_rules WHERE id=?", (rid,))
        db.commit(); db.close()
        return jsonify({'ok': True})
    d = request.json or {}
    db.execute("UPDATE fund_group_rules SET keyword=?, fund_group_id=? WHERE id=?",
               (d.get('keyword', ''), d.get('fund_group_id'), rid))
    db.commit(); db.close()
    return jsonify({'ok': True})


@app.route('/api/monthly-fund-budgets', methods=['GET', 'POST'])
def api_monthly_fund_budgets():
    db = get_db()
    if request.method == 'GET':
        year  = request.args.get('year')
        month = request.args.get('month')
        rows = db.execute(
            "SELECT b.*, g.name as fund_group_name FROM monthly_fund_budgets b "
            "LEFT JOIN fund_groups g ON b.fund_group_id = g.id "
            "WHERE b.year=? AND b.month=? ORDER BY g.sort_order, g.name",
            (year, int(month))
        ).fetchall()
        return jsonify(rows_to_list(rows))
    d = request.json or {}
    db.execute(
        "INSERT INTO monthly_fund_budgets (fund_group_id, year, month, budget_amount) VALUES (?,?,?,?) "
        "ON CONFLICT(fund_group_id, year, month) DO UPDATE SET budget_amount=excluded.budget_amount",
        (d.get('fund_group_id'), d.get('year'), d.get('month'), d.get('budget_amount', 0))
    )
    db.commit(); db.close()
    return jsonify({'ok': True}), 201


@app.route('/api/fund-summary')
def api_fund_summary():
    year  = request.args.get('year')
    month = request.args.get('month')
    db = get_db()
    # 자금 그룹별 실제 지출 집계
    actuals = db.execute(
        "SELECT g.id, g.name, g.sort_order, COALESCE(SUM(t.amount), 0) as actual "
        "FROM fund_groups g "
        "LEFT JOIN card_tx t ON t.fund_group_id = g.id "
        "  AND strftime('%Y', t.date) = ? AND strftime('%m', t.date) = ? "
        "GROUP BY g.id ORDER BY g.sort_order, g.name",
        (year, month.zfill(2))
    ).fetchall()
    # 월별 예산 조회
    budgets = {r['fund_group_id']: r['budget_amount'] for r in db.execute(
        "SELECT fund_group_id, budget_amount FROM monthly_fund_budgets WHERE year=? AND month=?",
        (year, int(month))
    ).fetchall()}
    db.close()
    result = []
    for row in actuals:
        result.append({
            'id': row['id'],
            'name': row['name'],
            'actual': row['actual'],
            'budget': budgets.get(row['id'], 0),
        })
    return jsonify(result)


@app.route('/api/card-tx/auto-fund-group', methods=['POST'])
def api_card_tx_auto_fund_group():
    data    = request.json or {}
    card_id = data.get('card_id')
    year    = data.get('year')
    month   = data.get('month')

    db = get_db()
    query  = "SELECT id, name FROM card_tx WHERE fund_group_locked = 0"
    params = []
    if card_id:
        query += " AND card_id = ?"; params.append(card_id)
    if year and month:
        query += " AND strftime('%Y', date) = ? AND strftime('%m', date) = ?"
        params += [year, month.zfill(2)]

    rows = db.execute(query, params).fetchall()
    updated = 0
    for row in rows:
        hint = _get_fund_group_hint(db, row['name'])
        if hint:
            db.execute("UPDATE card_tx SET fund_group_id=? WHERE id=?", (hint, row['id']))
            updated += 1
    db.commit()
    db.close()
    return jsonify({'ok': True, 'updated': updated})


@app.route('/api/fund-group-hint')
def api_fund_group_hint():
    name = request.args.get('name', '').strip()
    if not name:
        return jsonify({'fund_group_id': None})
    db = get_db()
    fund_group_id = _get_fund_group_hint(db, name)
    db.close()
    return jsonify({'fund_group_id': fund_group_id})


# ── API: AI 인사이트 (전문가 진단) ──────────────────────────
def _generate_asset_report(db, level=1):
    """전문가 진단을 위한 데이터 요약 보고서 생성 (Level 1: 현금 흐름 중심)"""
    today = date.today()
    this_year = today.year
    this_month = today.month
    
    report = f"### [자산 진단 데이터 보고서 - {today.isoformat()}]\n\n"
    
    if level == 1:
        # 1. 수입 내역 (최근 3개월)
        income_rows = db.execute("""
            SELECT strftime('%Y-%m', date) as ym, SUM(amount) as total
            FROM income 
            GROUP BY ym ORDER BY ym DESC LIMIT 3
        """).fetchall()
        report += "#### 1. 수입 현황 (최근 3개월)\n"
        for r in income_rows:
            report += f"- {r['ym']}: {r['total']:,}원\n"
        
        # 2. 지출 내역 (이번 달 카테고리별)
        expense_rows = db.execute("""
            SELECT category, SUM(amount) as total
            FROM budget 
            WHERE strftime('%Y', date) = ? AND strftime('%m', date) = ?
            GROUP BY category ORDER BY total DESC
        """, (str(this_year), str(this_month).zfill(2))).fetchall()
        report += f"\n#### 2. 이번 달({this_year}-{this_month}) 지출 상세\n"
        total_exp = 0
        for r in expense_rows:
            report += f"- {r['category'] or '미분류'}: {r['total']:,}원\n"
            total_exp += r['total']
        report += f"**총 지출 합계: {total_exp:,}원**\n"

    return report

@app.route('/api/ai-insight', methods=['GET'])
def api_ai_insight():
    level = int(request.args.get('level', 1))
    db = get_db()
    report_data = _generate_asset_report(db, level)
    
    # 실제 환경에서는 AI 모델을 호출하겠지만, 여기서는 데이터를 바탕으로 전문가 의견을 생성합니다.
    # 주거비 비중이 높은 실제 데이터를 반영한 예시 분석 결과
    analysis_text = f"""
### 📊 1단계: 현금 흐름 전문가 진단 결과

#### 1. 현 상태 진단
자네의 이번 달 현금 흐름을 보니, 주거비용이 지출의 상당 부분을 차지하고 있구만. 수입 대비 지출의 균형을 맞추는 것이 급선무라네.

#### 2. 위험 요소 및 개선점
- **고정비 과다:** 주거비나 통신비 등 숨만 쉬어도 나가는 돈의 비중이 너무 높네.
- **분류 체계:** 카테고리가 명확하지 않은 지출은 자산 관리의 적이라네.

#### 3. 전문가의 한 줄 평
> "자산의 건물은 현금 흐름이라는 기반 위에 지어지는 법. 기반이 흔들리면 성을 쌓을 수 없다네."
    """
    
    return jsonify({'ok': True, 'analysis': analysis_text, 'level': level})


# ── 동기화 페이지 ────────────────────────────────────────────
@app.route('/sync')
def sync_page():
    return render_template('sync.html')


SOURCE_FILES = [
    'app.py', 'database.py', 'requirements.txt',
    'templates/base.html', 'templates/dashboard.html',
    'templates/income.html', 'templates/budget.html',
    'templates/cards.html', 'templates/investments.html',
    'templates/realestate.html', 'templates/loans.html',
    'templates/pension.html', 'templates/goals.html',
    'templates/monthly.html', 'templates/sync.html',
    'templates/fund_management.html',
    '<link href="/static/css/style.css?v=12" rel="stylesheet">',
    'static/js/common.js', 'static/js/dashboard.js',
]

@app.route('/api/export-source')
def api_export_source():
    import base64
    files = {}
    base = os.path.dirname(os.path.abspath(__file__))
    for rel in SOURCE_FILES:
        path = os.path.join(base, rel.replace('/', os.sep))
        if os.path.exists(path):
            with open(path, 'rb') as f:
                files[rel] = base64.b64encode(f.read()).decode()
    return jsonify({'files': files})


@app.route('/api/compare-source', methods=['POST'])
def api_compare_source():
    import base64
    files = (request.json or {}).get('files', {})
    if not files:
        return jsonify({'error': '내용이 없습니다'}), 400
    base_dir = os.path.dirname(os.path.abspath(__file__))
    result = []
    for rel, b64 in files.items():
        dest = os.path.join(base_dir, rel.replace('/', os.sep))
        incoming = base64.b64decode(b64)
        if os.path.exists(dest):
            with open(dest, 'rb') as f:
                current = f.read()
            status = 'same' if incoming == current else 'changed'
        else:
            status = 'new'
        result.append({'file': rel, 'status': status, 'size': len(incoming)})
    return jsonify(result)


@app.route('/api/import-source', methods=['POST'])
def api_import_source():
    import base64
    files = (request.json or {}).get('files', {})
    if not files:
        return jsonify({'error': '내용이 없습니다'}), 400
    base = os.path.dirname(os.path.abspath(__file__))
    for rel, b64 in files.items():
        dest = os.path.join(base, rel.replace('/', os.sep))
        os.makedirs(os.path.dirname(dest), exist_ok=True)
        with open(dest, 'wb') as f:
            f.write(base64.b64decode(b64))
    return jsonify({'ok': True, 'count': len(files)})


@app.route('/api/export-text')
def api_export_text():
    return jsonify({'error': 'Cloud DB 환경에서는 지원되지 않습니다.'}), 400


@app.route('/api/import-text', methods=['POST'])
def api_import_text():
    return jsonify({'error': 'Cloud DB 환경에서는 지원되지 않습니다.'}), 400


if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=5000)
